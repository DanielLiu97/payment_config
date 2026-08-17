# -*- coding: utf-8 -*-
"""
生成支付配置校验Excel模板
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_template(output_path: str = "payment_config_template.xlsx"):
    """
    创建Excel模板文件
    """
    # 定义表头
    headers = [
        "国家",      # 列0 - 需要合并单元格
        "会员类型",  # 列1 - 需要合并单元格
        "价格类型",  # 列2 - 需要合并单元格
        "商品序号",  # 列3
        "周期",      # 列4
        "总价",      # 列5
        "均价",      # 列6
        "价格ID",    # 列7
        "商品ID",    # 列8
        "买赠周期",  # 列9
        "橱窗ID"     # 列10 - 格式：支付页：3553
    ]
    
    # 创建示例数据
    sample_data = [
        # 第一组：美国 Pro 原价
        ["US", "Pro", "原价", 1, "1年", 59.99, 4.99, "", "", "", "支付页：3553"],
        # 第二组：美国 Pro 折扣价
        ["US", "Pro", "折扣价", 1, "1年", 35.99, 1.99, "", "", "6个月", "支付页：3553"],
        # 第三组：美国 Pro 试用价
        ["US", "Pro", "试用价", 1, "1年", 59.99, 4.99, "", "", "", "支付页：3553"],
        # 第四组：美国 Pro 折扣价-3天试用
        ["US", "Pro", "折扣价-3天试用", 1, "1年", 35.99, 1.99, "", "", "6个月", "支付页：3553"],
    ]
    
    # 创建DataFrame
    df = pd.DataFrame(sample_data, columns=headers)
    
    # 保存为Excel
    df.to_excel(output_path, index=False, sheet_name="Sheet1")
    
    # 使用openpyxl美化格式
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 设置样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 设置表头样式
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # 设置列宽
    column_widths = {
        'A': 12,  # 国家
        'B': 15,  # 会员类型
        'C': 18,  # 价格类型
        'D': 10,  # 商品序号
        'E': 10,  # 周期
        'F': 12,  # 总价
        'G': 12,  # 均价
        'H': 12,  # 价格ID
        'I': 12,  # 商品ID
        'J': 15,  # 买赠周期
        'K': 20   # 橱窗ID
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 合并前3列的单元格（国家、会员类型、价格类型）
    # 示例：合并第2-5行的A列（国家）
    ws.merge_cells('A2:A5')  # 第一组数据
    ws.merge_cells('B2:B5')  # 会员类型
    ws.merge_cells('C2:C5')  # 价格类型
    
    # 设置合并单元格样式
    for row in range(2, 6):  # 第2-5行
        for col in range(1, 4):  # A-C列
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align
    
    # 设置数据行样式
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col <= 3:  # 前3列居中
                cell.alignment = center_align
            elif col in [4, 7, 8]:  # 商品序号、价格ID、商品ID居中
                cell.alignment = center_align
    
    # 添加说明行（在表头前插入）
    ws.insert_rows(1)
    ws.merge_cells('A1:K1')
    note_cell = ws.cell(row=1, column=1)
    note_cell.value = "说明：1. 前3列（国家、会员类型、价格类型）需要合并相同值的单元格；2. 橱窗ID格式：支付页：3553；3. 价格类型支持：原价、试用、试用价、折扣价、首优原价、折扣价-3天试用、挽回、挽回价、挽回试用、一次性原价、一次性折扣价、一次性挽回价、加购、加购试用、试用加购"
    note_cell.font = Font(size=9, italic=True)
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 调整第一行高度
    ws.row_dimensions[1].height = 40
    
    wb.save(output_path)
    print(f"Excel模板已生成: {output_path}")
    print("\n列说明：")
    print("列0: 国家（如：US、印度、T1等）")
    print("列1: 会员类型（如：Pro、Premium等）")
    print("列2: 价格类型（原价、折扣价、试用价等）")
    print("列3: 商品序号（从1开始，对应橱窗中的商品顺序）")
    print("列4: 周期（如：1年、2年、3个月等）")
    print("列5: 总价（美元，如：59.99）")
    print("列6: 均价（美元/月，如：4.99）")
    print("列7: 价格ID（程序会自动校验，可先留空）")
    print("列8: 商品ID（程序会自动校验，可先留空）")
    print("列9: 买赠周期（如：6个月，原价可不填）")
    print("列10: 橱窗ID（格式：支付页：3553，相同橱窗可只填第一行）")

if __name__ == "__main__":
    import os
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "payment_config_template.xlsx")
    create_excel_template(output_path)
