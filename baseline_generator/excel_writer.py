# -*- coding: utf-8 -*-
"""
基线配置 Excel 写出模块

将 generator 生成的行数据写入格式化的 Excel 文件，
格式与校验工具要求的模板完全一致。
"""
import os
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from baseline_generator.generator import BaselineRow, GenerateResult

HEADERS = [
    "国家", "会员类型", "价格类型", "商品序号",
    "周期", "总价", "均价", "价格ID",
    "商品ID", "买赠周期", "橱窗ID",
    "体验价价格", "体验价周期",
]

COLUMN_WIDTHS = {
    "A": 10, "B": 18, "C": 18, "D": 10,
    "E": 10, "F": 12, "G": 12, "H": 12,
    "I": 12, "J": 15, "K": 18,
    "L": 14, "M": 14,
}


def _apply_styles(ws, data_row_count: int):
    """应用表头样式、列宽、边框"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    for row in range(2, data_row_count + 2):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col <= 3 or col in (4, 8, 9):
                cell.alignment = center_align


def _merge_repeated_cells(ws, data_row_count: int):
    """合并 A~C 列中连续相同值的单元格（国家、会员类型、价格类型）"""
    for col_idx in range(1, 4):
        start_row = 2
        while start_row <= data_row_count + 1:
            current_val = ws.cell(row=start_row, column=col_idx).value
            end_row = start_row
            while end_row + 1 <= data_row_count + 1:
                next_val = ws.cell(row=end_row + 1, column=col_idx).value
                if next_val == current_val:
                    end_row += 1
                else:
                    break
            if end_row > start_row:
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                merged_cell = ws.cell(row=start_row, column=col_idx)
                merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            start_row = end_row + 1


def write_baseline_excel(result: GenerateResult, output_path: str) -> str:
    """
    将基线数据写入 Excel 文件。

    Args:
        result: GenerateResult 对象
        output_path: 输出文件路径

    Returns:
        实际写入的文件路径
    """
    rows_data = [row.to_list() for row in result.rows]

    df = pd.DataFrame(rows_data, columns=HEADERS)
    df.to_excel(output_path, index=False, sheet_name="Sheet1")

    wb = load_workbook(output_path)
    ws = wb.active
    data_row_count = len(rows_data)

    _apply_styles(ws, data_row_count)
    _merge_repeated_cells(ws, data_row_count)

    # 添加摘要信息到第二个 sheet
    summary_ws = wb.create_sheet("生成摘要")
    summary_items = [
        ("生成参数", ""),
        ("橱窗ID", ", ".join(str(s) for s in result.showcase_ids)),
        ("平台", result.platform),
        ("环境", result.env),
        ("国家", result.country),
        ("", ""),
        ("生成结果", ""),
        ("总行数", str(len(result.rows))),
        ("警告数", str(len(result.warnings))),
    ]
    for warn in result.warnings:
        summary_items.append(("警告", warn))

    summary_items.extend([
        ("", ""),
        ("使用说明", ""),
        ("1", "本文件为基线配置，数据来自橱窗接口实时查询"),
        ("2", "请基于此文件调整配置后，上传至校验工具进行校验"),
        ("3", "首优折扣率(J列)、部分条件商品可能未包含，请手动补充"),
    ])

    for i, (key, val) in enumerate(summary_items, 1):
        summary_ws.cell(row=i, column=1, value=key)
        summary_ws.cell(row=i, column=2, value=val)
        if key in ("生成参数", "生成结果", "使用说明"):
            summary_ws.cell(row=i, column=1).font = Font(bold=True, size=12)

    summary_ws.column_dimensions["A"].width = 16
    summary_ws.column_dimensions["B"].width = 60

    wb.save(output_path)
    return output_path
